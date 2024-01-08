import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

def csv_to_xlsx(file_name, path, Id):
    df = pd.read_csv(file_name, delimiter=';')
    df.to_excel(f'{path}/ready_mate_data.xlsx', index=False)
    excel_file = f'{path}/ready_mate_data.xlsx'
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    alignment = Alignment(horizontal='left')
    font_bold = Font(bold=True)
    
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 180
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment
            
    for cell in ws['A']:
        cell.font = font_bold
        
    for cell in ws['B']:
        cell.font = Font()
        
    links_column = 2
    start_row = 2
    
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=links_column)
        if cell.value and str(cell.value).startswith('http'):
            cell.style = 'Hyperlink'
            cell.hyperlink = cell.value
            
    wb.save(excel_file)
    os.remove(f'{path}/ready_mate_data.csv')
    print(f'[metabase] The table for the call {Id} has been formed!')