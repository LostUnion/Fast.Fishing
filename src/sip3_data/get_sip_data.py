import csv
import time
import requests
from datetime import datetime
from src.sessions import user_agent
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle

def time_format():
    time_data = datetime.strptime(datetime.now().strftime("%Y/%m/%d %H:%M:%S"), "%Y/%m/%d %H:%M:%S")
    terminated_at = str(time.mktime(time_data.timetuple()))[:-2]
    return terminated_at

def sip(external_id, path):
    try:
        URL = 'http://172.16.4.19/api/search'

        query = f"sip.call_id={external_id}"

        terminated_at = time_format()

        payload = {
                "created_at" : 0,
                "limit" : 50,
                "query" : query,
                "terminated_at" : terminated_at+'000',
            }

        headers = {
                    "accept" : "application/json",
                    "content-type" : "application/json",
                    "User-Agent": user_agent
                }

        res = requests.post(URL, headers=headers, json=payload)
        print(f'[sip3][{res.status_code}] Connecting to SIP3')
        print('[sip3] Getting information from SIP3\n')
        
        link_sip = f'http://172.16.4.19/#/search/advanced?created_at={res.json()[0]['created_at']}&terminated_at={res.json()[0]['terminated_at']}&query=sip.call_id={res.json()[0]['call_id'][0]}'
        
        s = ['SIP3', link_sip]
        ready_mate_data = f'{path}/ready_mate_data.xlsx'
        
        wb = load_workbook(ready_mate_data)
        sheet = wb.active
        sheet.append(s)
        
        row_number = sheet.max_row
        column_number = 1
        
        cell = sheet.cell(row=row_number, column=column_number)
        cell.font = Font(bold=True)

        column_number = 2

        cell = sheet.cell(row=row_number, column=column_number)
        cell.font = Font(underline='single')

        hyperlink_style = NamedStyle(name='hyperlink')
        hyperlink_style.font = Font(underline='single')
        wb.add_named_style(hyperlink_style)
        
        cell.style = 'Hyperlink'
        cell.hyperlink = cell.value
        
        wb.save(ready_mate_data)
    except:
        print('[sip3] The data was not found, request the current call ID.')