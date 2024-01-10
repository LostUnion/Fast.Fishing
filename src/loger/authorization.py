import requests
import base64
from datetime import datetime
from src.database.database import db_auth, get_login, get_password, delete_login_and_password
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from urllib.parse import quote

def times(Started_At_Updates, Finished_At_Updates):
    started_at = datetime.strptime(Started_At_Updates, "%Y/%m/%d %H:%M")
    finished_at = datetime.strptime(Finished_At_Updates, "%Y/%m/%d %H:%M")

    formatted_started_date_full = started_at.strftime("%Y-%m-%d %H:%M:00")
    formatted_finished_date_full = finished_at.strftime("%Y-%m-%d %H:%M:59")

    formatted_started_date_link = started_at.strftime("%Y-%m-%d")
    formatted_finished_date_link = finished_at.strftime("%Y-%m-%d")

    formatted_started_time_link = started_at.strftime("%H:%M:00")
    formatted_finished_time_link = finished_at.strftime("%H:%M:59")

    formatted_started_time_link = quote(formatted_started_time_link)
    formatted_finished_time_link = quote(formatted_finished_time_link)

    formatted_started_date_link = f'{formatted_started_date_link}+{formatted_started_time_link}%2B0000'
    formatted_finished_date_link = f'{formatted_finished_date_link}+{formatted_finished_time_link}%2B0000'

    return formatted_started_date_full, formatted_finished_date_full, formatted_started_date_link, formatted_finished_date_link


def get_token():
    db_auth('loggers')
    login = get_login('loggers')
    password = get_password('loggers')

    token = str(base64.b64encode(f'{login}:{password}'.encode('ascii')))[2:-1]
    return token

def get_info(call_id, Started_At_Updates, Finished_At_Updates, path):
    while True:

        token = get_token()

        formatted_started_date_full, formatted_finished_date_full, formatted_started_date_link, formatted_finished_date_link = times(Started_At_Updates, Finished_At_Updates)

        URL = f'https://analyse.twin24.ai/logs'

        payload = {
            'timeSort' : 'desc',
            'payload' : f'{call_id}',
            'title' : '',
            'traceId' : '',
            'requestId' : '',
            'from' : f'{formatted_started_date_full}',
            'to' : f'{formatted_finished_date_full}',
            'minexectime' : '',
            'maxexectime' : ''
        }

        headers = {
                    'Authorization': f'Basic {token}'
                }
        
        res = requests.get(URL, headers=headers, params=payload)

        if res.status_code == 200:
            link = f'https://analyse.twin24.ai/logs?timeSort={payload['timeSort']}&payload={payload['payload']}&title=&traceId=&requestId=&from={formatted_started_date_link}&to={formatted_finished_date_link}&minexectime=&maxexectime='

            s = ['LOGS', link]
            ready_mate_data = f'{path}/ready_mate_data.xlsx'

            wb = load_workbook(ready_mate_data)
            sheet = wb.active
            sheet.append(s)

            row_number = sheet.max_row
            
            style_exists = False
            for style in wb._named_styles:
                if style.name == 'hyperlink':
                    style_exists = True
                    break
                
            if not style_exists:
                hyperlink_style = NamedStyle(name='hyperlink')
                hyperlink_style.font = Font(underline='single')
                wb._named_styles.append(hyperlink_style)

            column_number = 1

            cell = sheet.cell(row=row_number, column=column_number)
            cell.font = Font(bold=True)

            column_number = 2

            cell = sheet.cell(row=row_number, column=column_number)
            cell.font = Font(underline='single')

            cell.style = 'hyperlink'
            cell.hyperlink = cell.value

            wb.save(ready_mate_data)
            break
        # elif str(res.status_code)[0] == "4" :
        #     print('Invalid username or password')   
        #     delete_login_and_password('loggers')
        # elif str(res.status_code)[0] == "5":
        #     print('Server error')
        #     delete_login_and_password('loggers')
